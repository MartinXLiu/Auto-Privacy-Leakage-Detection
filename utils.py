import zipfile
def decompress_zip(zip_path, dir_path):
    """
    zip 文件解压
    :param zip_path: zip 压缩包路径
    :param dir_path: 解压路径
    :return:
    """
    f = zipfile.ZipFile(zip_path, 'r')  # 压缩文件位置
    for file in f.namelist():
        f.extract(file, dir_path)  # 解压位置
    f.close()



import os
import rarfile
def decompress_rar(rar_file_name, dir_name):
    """
    rar 文件解压
    :param rar_file_name: rar 文件路径
    :param dir_name: 文件解压目录
    :return:
    """
    # 创建 rar 对象
    rar_obj = rarfile.RarFile(rar_file_name)
    # 目录切换
    if not os.path.exists(dir_name):
        os.mkdir(dir_name)
    # Extract all files into current directory.
    rar_obj.extractall(dir_name)
    # 关闭
    rar_obj.close()
# rar解压报错：rarfile.BadRarFile: Failed the read enough data: req=16384 got=52-->解决办法：https://blog.csdn.net/xrinosvip/article/details/120396624
# 可以借鉴这个代码进行文件解压缩相关函数方法的整合：https://blog.csdn.net/m0_62026333/article/details/127908757?ops_request_misc=%257B%2522request%255Fid%2522%253A%2522169203110116800225523252%2522%252C%2522scm%2522%253A%252220140713.130102334..%2522%257D&request_id=169203110116800225523252&biz_id=0&utm_medium=distribute.pc_search_result.none-task-blog-2~all~sobaiduend~default-2-127908757-null-null.142^v92^controlT0_1&utm_term=rarfile.BadRarFile%3A%20Failed%20the%20read%20enough%20data%3A%20req%3D16384%20got%3D52&spm=1018.2226.3001.4187



def read_txt(txt_path, output_path):
    """
    读取文本文件
    :param txt_path: 文本文件路径
    :param output_path: 输出文件路径
    :return:
    """
    with open(txt_path, "r", encoding="utf-8") as txt_file:
        data = txt_file.read()
        with open(output_path, "a", encoding="utf-8") as output_file:
            output_file.write(data)



import win32com.client as wc
def doc2docx(input_path, output_path):
    """
    把doc文件另存为docx文件
    :param input_path: 输入doc文件路径（完整路径）
    :param output_path: 输出docx文件路径（完整路径）
    :return:
    """
    word = wc.Dispatch('Word.Application')
    doc = word.Documents.Open(input_path)
    doc.SaveAs(output_path, 12)
    doc.Close()
    word.Quit()
# 路径报错：pywintypes.com_error-->解决办法：https://blog.csdn.net/qq_38390215/article/details/131195395?ops_request_misc=&request_id=&biz_id=102&utm_term=pywintypes.com_error:&utm_medium=distribute.pc_search_result.none-task-blog-2~all~sobaiduweb~default-2-131195395.142^v92^controlT0_1&spm=1018.2226.3001.4187



import docx
def read_docx(docx_path, output_path):
    """
    读取docx文件中的文字
    :param docx_path: docx文件路径（完整路径）
    :param output_path: 输出文件路径
    :return:
    """
    docx_file = docx.Document(docx_path)
    with open(output_path, "a", encoding="utf-8") as output_file:
        for para in docx_file.paragraphs:
            output_file.write(para.text + '\n')



import win32com.client as wc
def wps_docx(input_path, output_path):
    """
    把wps文件另存为docx文件
    :param input_path: 输入wps文件路径（完整路径）
    :param output_path: 输出docx文件路径（完整路径）
    :return:
    """
    kwps = wc.Dispatch("Kwps.Application")
    wps = kwps.Documents.Open(input_path)
    wps.SaveAs(output_path, 12)
    wps.Close()
    kwps.Quit()



import openpyxl
def read_xlsx(xlsx_path, output_path):
    """
    读取xlsx文件中的文字
    :param xlsx_path: xlsx文件路径
    :param output_path: 输出文件路径
    :return:
    """
    work_book = openpyxl.load_workbook(xlsx_path)
    with open(output_path, "a", encoding="utf-8") as output_file:
        for sheet_name in work_book.sheetnames:
            output_file.write('\n')
            work_sheet = work_book[sheet_name]
            for ws_row in range(1, work_sheet.max_row + 1):
                for ws_column in range(1, work_sheet.max_column + 1):
                    output_file.write(str(work_sheet.cell(row=ws_row, column=ws_column).value) + ' ')



import win32com.client as wc
def et_xlsx(input_path, output_path):
    """
    把et文件另存为xlsx文件
    :param input_path: 输入et文件路径（完整路径）
    :param output_path: 输出xlsx文件路径（完整路径）
    :return:
    """
    ket = wc.Dispatch("Ket.Application")
    et = ket.Workbooks.Open(input_path)
    et.SaveAs(output_path, 12)
    et.Close()
    ket.Quit()



if __name__ == '__main__':
    pass
    # decompress_zip('file_test/python_fasts3-main.zip', 'file_test/python_fasts3-main')
    # decompress_rar('file_test/题目1：富文本敏感信息泄露检测.rar', 'file_test/题目1：富文本敏感信息泄露检测')
    # read_txt('file_test/赛题材料 - 副本/token', 'file_test/output.txt')
    # doc2docx(r'D:\OneDrive\CanWorkSpace\Code\PycharmSpace\Auto-Privacy-Leakage-Detection\file_test\赛题材料 - 副本\Android手机VPN安装指南.doc', r'D:\OneDrive\CanWorkSpace\Code\PycharmSpace\Auto-Privacy-Leakage-Detection\file_test\赛题材料 - 副本\Android手机VPN安装指南.docx')
    # read_docx(r'D:\OneDrive\CanWorkSpace\Code\PycharmSpace\Auto-Privacy-Leakage-Detection\file_test\赛题材料 - 副本\Android手机VPN安装指南.docx', 'file_test/output.txt')
    # wps_docx(r'D:\OneDrive\CanWorkSpace\Code\PycharmSpace\Auto-Privacy-Leakage-Detection\file_test\赛题材料 - 副本\Android手机VPN安装指南.wps', r'D:\OneDrive\CanWorkSpace\Code\PycharmSpace\Auto-Privacy-Leakage-Detection\file_test\赛题材料 - 副本\Android手机VPN安装指南.docx')
    # read_xlsx('file_test/赛题材料 - 副本/资产梳理.xlsx', 'file_test/output.txt')
    # et_xlsx(r'D:\OneDrive\CanWorkSpace\Code\PycharmSpace\Auto-Privacy-Leakage-Detection\file_test\赛题材料 - 副本\资产梳理.et', r'D:\OneDrive\CanWorkSpace\Code\PycharmSpace\Auto-Privacy-Leakage-Detection\file_test\赛题材料 - 副本\资产梳理.xlsx')